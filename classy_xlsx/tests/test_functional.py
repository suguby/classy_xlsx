# -*- coding: utf-8 -*-
import os
import tempfile
from unittest import TestCase

from os import unlink

import datetime

from openpyxl import load_workbook


from classy_xlsx.columns import (
    FloatXlsxColumn, IntegerXlsxColumn, PercentXlsxColumn, RatioColumn, WeightedAverageColumn,
    TextXlsxColumn, UnicodeXlsxColumn,
    DateXlsxColumn, DateTimeXlsxColumn
)
from classy_xlsx.core import Bunch
from classy_xlsx.regions import XlsxTable
from classy_xlsx.workbook import XlsxWorkbook
from classy_xlsx.worksheet import XlsxSheet


class TestRatioColumn(RatioColumn):
    dividend = 'int_col'
    divisor = 'float_col'


class TestWeightedAverageColumn(WeightedAverageColumn):
    divisor = 'int_col'


class TestXlsxTable(XlsxTable):
    int_col = IntegerXlsxColumn(title=u'Первая')
    float_col = FloatXlsxColumn(title=u'Вторая')
    percent_col = PercentXlsxColumn()
    ratio_col = TestRatioColumn()
    wa_col = TestWeightedAverageColumn()
    text_col = TextXlsxColumn()
    uni_col = UnicodeXlsxColumn()
    date_col = DateXlsxColumn()
    datetime_col = DateTimeXlsxColumn()

    def get_queryset(self):
        return [
            Bunch(
                int_col=1,
                float_col=10.1,
                percent_col=0.3,
                wa_col=100,
                text_col='some text',
                uni_col=u'какой-то текст',
                date_col=datetime.date(year=2016, month=01, day=01),
                datetime_col=datetime.datetime(year=2016, month=01, day=01, hour=12),
            ),
            Bunch(
                int_col=2,
                float_col=20.2,
                percent_col=0.5,
                wa_col=50,
                text_col='some text 2',
                uni_col=u'какой-то текст еще',
                date_col=datetime.date(year=2016, month=06, day=11),
                datetime_col=datetime.datetime(year=2016, month=11, day=11, hour=12),
            ),
        ]


class TestXlsxSheet(XlsxSheet):
    table = TestXlsxTable()


class TestXlsxWorkbook(XlsxWorkbook):
    sheet = TestXlsxSheet()

result_sheet_data = {
    'A1': Bunch(value=u'Первая'),
    'B1': Bunch(value=u'Вторая'),
    'C1': Bunch(value='Column3'),
    'D1': Bunch(value='Column4'),

    'A2': Bunch(value=1),
    'B2': Bunch(value=10.1),
    'C2': Bunch(value=0.3),
    'D2': Bunch(value=0.099009900990099),
    'E2': Bunch(value=100),
    'F2': Bunch(value='some text'),
    'G2': Bunch(value=u'какой-то текст'),
    'H2': Bunch(value='01.01.2016'),
    'I2': Bunch(value='01.01.2016 12:00:00'),

    'A3': Bunch(value=2),
    'B3': Bunch(value=20.2),
    'C3': Bunch(value=0.5),
    'D3': Bunch(value=0.099009900990099),
    'E3': Bunch(value=50),
    'F3': Bunch(value='some text 2'),
    'G3': Bunch(value=u'какой-то текст еще'),
    'H3': Bunch(value='11.06.2016'),
    'I3': Bunch(value='11.11.2016 12:00:00'),

}

class ContextTest(TestCase):

    def setUp(self):
        ff, self.out_file_name = tempfile.mkstemp(suffix='.xlsx')
        os.close(ff)
        self.wb = TestXlsxWorkbook(file_name=self.out_file_name)

    def test_all(self):
        self.wb.make_report()
        wb2 = load_workbook(self.wb.file_name)
        self.assertIn('sheet', wb2.get_sheet_names())
        sheet = wb2.get_sheet_by_name('sheet')
        for k, v in result_sheet_data.iteritems():
            self.assertEquals(sheet[k].value, v.value)


        # def tearDown(self):
    #     unlink(self.out_file_name)
